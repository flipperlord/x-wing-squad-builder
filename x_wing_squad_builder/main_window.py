import logging
import os
from pathlib import Path
from PySide6 import QtWidgets, QtCore, QtGui

from x_wing_squad_builder.definition_form import DefinitionForm
from x_wing_squad_builder.upgrade_form import UpgradeForm
from x_wing_squad_builder.viewer import Viewer
from .settings import Settings
from .worker import Worker
from .root_logger_handler import RootLoggerHandler
from .ui import DarkPalette, IconPath
from .ui.main_window_ui import Ui_MainWindow
from .about_window import AboutWindow
from .settings_window import SettingsWindow

from .model import XWing, PilotEquip, Squad, Upgrades

from .utils_pyside import (image_path_to_qpixmap, populate_list_widget, update_action_layout,
                           update_upgrade_slot_layout, treewidget_item_is_top_level,
                           )
from .utils import (get_upgrade_slot_from_list_item_text, gui_text_decode, prettify_name, gui_text_encode,
                    get_pilot_name_from_list_item_text, get_upgrade_name_from_list_item_text)

from pathlib import Path

from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class MainWindow(QtWidgets.QMainWindow):
    """Main Window"""

    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        self.application_name = "X-Wing Squad Builder"
        self.organization_name = "ryanlenardryanlenard, Inc."

        self.settings = Settings()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowTitle(self.application_name)

        self.configure_logging()
        logging.info(f"Welcome to {self.application_name}")

        # Setup threadpool
        self.threadpool = QtCore.QThreadPool()

        # To create an async worker
        # worker = Worker(func)
        # self.threadpool.start(worker)

        # Setup Windows
        self.about_window = AboutWindow()
        self.settings_window = SettingsWindow()
        self.settings_window.ui.theme_combo_box.currentTextChanged.connect(
            self.check_theme)
        self.settings_window.saved_signal.connect(self.reload_data)

        # Add widgets with icon paths here to be inverted on a theme change.
        self.widgets_with_icons = {
            self.ui.action_open_settings_window: IconPath.SETTINGS.value,
            self.settings_window.ui.select_log_file_directory_push_button: IconPath.OPEN.value,
            self.ui.menu_excel: IconPath.EXCEL.value
        }

        self.check_theme()

        # Connect signals and slots
        self.ui.action_about.triggered.connect(self.about_window.show)
        self.ui.action_exit.triggered.connect(self.close)
        self.ui.action_open_settings_window.triggered.connect(
            self.handle_show_settings_window)
        self.ui.action_export_as_excel.triggered.connect(self.export_excel)
        self.ui.action_import_excel.triggered.connect(self.import_excel)

        self.ui.faction_list_widget.itemSelectionChanged.connect(self.update_faction)
        self.ui.ship_list_widget.itemSelectionChanged.connect(self.update_ship)
        self.ui.pilot_list_widget.itemSelectionChanged.connect(self.update_pilot)
        self.ui.pilot_list_widget.enter_signal.connect(self.handle_equip_pilot)
        self.ui.pilot_list_widget.itemDoubleClicked.connect(self.handle_equip_pilot)
        self.ui.upgrade_list_widget.itemSelectionChanged.connect(self.update_upgrade)
        self.ui.upgrade_list_widget.enter_signal.connect(self.handle_equip_upgrade)
        self.ui.upgrade_list_widget.itemDoubleClicked.connect(self.handle_equip_upgrade)

        # Initialize Factions
        self.file_path = self.data_dir / "definition.json"

        # Initialize widgets
        self.ui.ship_name_label.clear()
        self.ui.base_label.clear()
        self.ui.initiative_label.clear()
        self.ui.points_label.clear()
        self.ui.maneuver_image_label.clear()
        self.ui.total_pilot_cost_label.clear()
        self.ui.total_upgrade_cost.clear()
        self.ui.total_cost_label.clear()
        self.ui.pilot_keyword_label.clear()

        # Set up definition form for adding to definition file.
        self.definition_form = self.initialize_definition_form()

        # Set up upgrade form for adding to definition file.
        self.ui.action_upgrade_form.triggered.connect(
            self.handle_show_upgrade_form)
        self.upgrade_form = self.initialize_upgrade_form()

        self.ui.action_reload_data.triggered.connect(self.reload_data)

        self.ui.equip_pilot_push_button.clicked.connect(self.handle_equip_pilot)
        self.ui.unequip_pilot_push_button.clicked.connect(self.unequip_pilot)
        self.ui.copy_pilot_push_button.clicked.connect(self.handle_copy_pilot)
        self.ui.equip_upgrade_push_button.clicked.connect(self.handle_equip_upgrade)
        self.ui.unequip_upgrade_push_button.clicked.connect(
            self.unequip_upgrade)
        self.ui.squad_tree_widget.itemSelectionChanged.connect(self.handle_squad_click)
        self.ui.squad_tree_widget.itemDoubleClicked.connect(self.handle_squad_double_click)

        self.squad = Squad()

        self.xwing = XWing.launch_xwing_data(self.file_path)
        self.upgrades = Upgrades(self.xwing.upgrades)

        # Set up upgrade viewer
        self.viewer = self.initialize_card_viewer()

        self.reload_data()

        self.update_costs()

        self.showMaximized()

    def handle_squad_timer(self):
        self.viewer.populate_squad_viewer(self.squad)

    def initialize_definition_form(self):
        definition_form = DefinitionForm(self.file_path)
        self.ui.action_definition_form.triggered.connect(definition_form.show)
        definition_form.update_signal.connect(self.reload_data)
        definition_form.form_closed_signal.connect(self.handle_form_closed)
        return definition_form

    def initialize_upgrade_form(self):
        """
        this creates an upgrade form object and assigns relevant signals/slots
        motivation for this is so we can easily reset the form when editing upgrades
        """
        upgrade_form = UpgradeForm(self.file_path)
        upgrade_form.update_signal.connect(self.handle_new_upgrade_data)
        upgrade_form.update_form_closed_signal.connect(self.handle_form_closed)
        return upgrade_form

    def initialize_card_viewer(self):
        viewer = Viewer(self.xwing, self.upgrades, self.upgrade_slots_dir,
                        self.upgrades_dir, self.factions_dir, self.ship_icons_dir, self.pilots_dir)
        self.ui.action_viewer.triggered.connect(viewer.show)
        viewer.upgrade_edit_signal.connect(self.edit_upgrade)
        viewer.pilot_edit_signal.connect(self.edit_pilot)
        return viewer

    def handle_form_closed(self):
        self.definition_form.edit_mode = False
        self.definition_form.edit_pilot_name = None
        self.definition_form.edit_upgrade_name = None
        self.definition_form.edit_ship_name = None

    def handle_show_upgrade_form(self):
        self.upgrade_form.ui.upgrade_name_line_edit.setReadOnly(False)
        self.upgrade_form.show()

    def edit_upgrade(self, upgrade_name):
        # This resets the form to defaults
        self.upgrade_form = self.initialize_upgrade_form()
        # self.upgrade_form.ui.upgrade_name_line_edit.setReadOnly(True)
        self.definition_form.edit_mode = True
        self.definition_form.edit_upgrade_name = upgrade_name
        upgrade_dict = self.upgrades.get_upgrade(upgrade_name)
        self.upgrade_form.populate_upgrade_form(upgrade_dict)
        self.upgrade_form.show()

    def edit_pilot(self, pilot_name, ship_name, faction_name):
        # This resets the form to defaults
        self.definition_form = self.initialize_definition_form()
        self.definition_form.edit_mode = True
        self.definition_form.edit_pilot_name = pilot_name
        self.definition_form.edit_ship_name = ship_name
        ship = self.xwing.get_ship(faction_name, ship_name)
        pilot = self.xwing.get_pilot(faction_name, ship_name, pilot_name)
        self.definition_form.populate_definition_form(ship, pilot)
        self.definition_form.show()

    def reload_data(self):
        self.definition_form.load_data()
        self.ui.ship_list_widget.clear()
        self.ui.pilot_list_widget.clear()
        self.ui.faction_list_widget.clear()
        self.ui.upgrade_list_widget.clear()
        self.xwing = XWing.launch_xwing_data(self.file_path)
        self.upgrades = Upgrades(self.xwing.upgrades)
        self.viewer.upgrades = self.upgrades
        self.viewer.xwing = self.xwing
        self.viewer.populate_upgrade_viewer()
        self.viewer.populate_pilot_viewer()
        # self.viewer = self.initialize_card_viewer()
        faction_names = [prettify_name(faction)
                         for faction in self.xwing.faction_names]
        populate_list_widget(
            faction_names, self.ui.faction_list_widget, self.factions_dir)
        # populate_list_widget(
        #     self.upgrades.all_upgrades_for_gui, self.ui.upgrade_list_widget)

    def handle_squad_click(self):
        item = self.squad_tree_selection
        if item is None:
            return
        self.ui.upgrade_list_widget.clear()

        # If you click on a pilot...
        if treewidget_item_is_top_level(item):
            pilot_data = self.squad.get_pilot_data(item)
            filtered_for_gui = self.upgrades.filtered_upgrades_for_gui(
                pilot_data.filtered_upgrades)
        # If you click on an upgrade slot...
        else:
            pilot_data = self.squad.get_pilot_data(item.parent())
            upgrade_slot = get_upgrade_slot_from_list_item_text(item.text(0))
            filtered_upgrades = self.upgrades.filtered_upgrades_by_pilot_and_slot(
                pilot_data, upgrade_slot)
            filtered_for_gui = self.upgrades.filtered_upgrades_for_gui(
                filtered_upgrades)

        populate_list_widget(filtered_for_gui, self.ui.upgrade_list_widget)
        self.pilot_image_label = pilot_data.pilot_name

    def handle_squad_double_click(self):
        item = self.squad_tree_selection
        if item is None:
            return

        if treewidget_item_is_top_level(item):
            self.unequip_pilot()
        else:
            self.unequip_upgrade()

    def refresh_squad_upgrade_slots(self, parent_select_item: QtWidgets.QTreeWidgetItem = None, select_item: QtWidgets.QTreeWidgetItem = None, auto_include_bypass = True):
        """rebuilds the squad list widget.  pass in args if you wish to set selection to the same prior to refresh"""
        for item, pilot_data in self.squad.squad_dict.items():
            # first clear the list
            for i in reversed(range(item.childCount())):
                item.removeChild(item.child(i))
            # Add all child slots to the pilot item
            for slot in pilot_data.upgrade_slots:
                child = QtWidgets.QTreeWidgetItem([prettify_name(slot)])
                pixmap = image_path_to_qpixmap(
                    self.upgrade_slots_dir / f"{slot}.png")
                child.setIcon(0, pixmap)
                item.addChild(child)
                # sets the selection to the item clicked before the refresh
                if parent_select_item == item and select_item.text(0).lower() == slot:
                    self.ui.squad_tree_widget.clearSelection()
                    child.setSelected(True)

            # Update available upgrades based on squad
            pilot_data.filtered_upgrades = self.upgrades.filtered_upgrades_by_pilot(
                pilot_data, self.squad)

            if not auto_include_bypass:
                for upgrade in pilot_data.filtered_upgrades:
                    auto_include = upgrade.get("autoinclude", "False") == "True"
                    if auto_include:
                        slots = self.upgrades.get_upgrade_slots(upgrade)
                        equipped = pilot_data.equip_upgrade(slots, upgrade.get("name"), upgrade["cost"], upgrade)
                        if equipped:
                            logging.info(f"{prettify_name(upgrade['name'])} equipped automatically to {prettify_name(pilot_data.pilot_name)}")

            # Now color in equipped upgrades
            self.color_in_equipped_upgrades(item, pilot_data)
            self.ui.squad_tree_widget.resizeColumnToContents(0)

    def color_in_equipped_upgrades(self, parent_item: QtWidgets.QTreeWidgetItem, pilot_data: PilotEquip):
        for upgrade in pilot_data.equipped_upgrades:
            required_slots = upgrade.attributes['upgrade_slot_types']
            for slot in required_slots:
                # Find an available item
                for child_idx in range(parent_item.childCount()):
                    child_item = parent_item.child(child_idx)
                    potential_slot = child_item.text(0).lower()
                    gui_equipped = child_item.text(1)
                    # empty string means nothing equipped here
                    if len(gui_equipped) == 0 and potential_slot == slot:
                        # Update icon to green (for equipped)
                        pixmap = image_path_to_qpixmap(
                            self.upgrade_slots_dir / f"{slot}.png", color="green")
                        child_item.setIcon(0, pixmap)
                        # update item column to the name
                        child_item.setText(
                            1, f'{prettify_name(upgrade.name)} ({upgrade.cost})')
                        break

    def handle_copy_pilot(self):
        if self.squad_tree_selection is None or not treewidget_item_is_top_level(self.squad_tree_selection):
            logging.info(
                "No pilot selected in squad tree - select an equipped pilot and try again.")
            return
        top_level_idx = self.ui.squad_tree_widget.indexOfTopLevelItem(
            self.squad_tree_selection)
        item = self.ui.squad_tree_widget.topLevelItem(top_level_idx)
        pilot_data = self.squad.get_pilot_data(item)
        equipped_item = self.equip_pilot(pilot_data.faction_name, pilot_data.ship_name, pilot_data.pilot_name)
        if equipped_item is not None:
            new_pilot_data = self.squad.get_pilot_data(equipped_item)
            for upgrade in pilot_data.equipped_upgrades:
                if upgrade.name in [u['name'] for u in new_pilot_data.filtered_upgrades]:
                    self.equip_upgrade(upgrade.name, new_pilot_data)
                else:
                    logging.info(f"Unable to equip {prettify_name(upgrade.name)}")
            logging.info(f"{prettify_name(pilot_data.pilot_name)} successfully copied.")
        else:
            return

    def handle_equip_pilot(self):
        if not self.pilot_name_selected:
            logging.info("No pilot selected - select a pilot and try again.")
            return
        faction_name = self.faction_selected
        ship_name = self.ship_selected_encoded
        if faction_name is None or ship_name is None:
            return
        self.equip_pilot(faction_name, ship_name, self.pilot_name_selected)

    def equip_pilot(self, faction_name: str, ship_name: str, pilot_name: str) -> Optional[QtWidgets.QTreeWidgetItem]:
        pilot = self.xwing.get_pilot(
            faction_name, ship_name, pilot_name)
        ship = self.xwing.get_ship(faction_name, ship_name)
        pilot_data = PilotEquip(ship, pilot)
        # TODO: Turn this formatting into a function as it's also used by the ship module
        item = QtWidgets.QTreeWidgetItem([f"({pilot_data.initiative}) {prettify_name(pilot_name)} ({pilot_data.cost})"])
        added = self.squad.add_pilot(item, pilot_data)
        if not added:
            return None

        self.ui.squad_tree_widget.insertTopLevelItem(
            self.squad_tree_bottom_index, item)
        self.refresh_squad_upgrade_slots(auto_include_bypass=False)
        # self.ui.squad_tree_widget.resizeColumnToContents(0)
        self.ui.squad_tree_widget.expandAll()
        self.update_costs()
        self.viewer.populate_squad_viewer(self.squad)
        return item

    def unequip_pilot(self):
        if self.squad_tree_selection is None:
            logging.info(
                "No pilot selected in squad tree - select a squad tree item and try again.")
        top_level_idx = self.ui.squad_tree_widget.indexOfTopLevelItem(
            self.squad_tree_selection)
        item = self.ui.squad_tree_widget.topLevelItem(top_level_idx)
        removed = self.squad.remove_pilot(item)
        if removed:
            self.ui.squad_tree_widget.takeTopLevelItem(top_level_idx)
            self.refresh_squad_upgrade_slots()
            self.update_costs()
            self.viewer.populate_squad_viewer(self.squad)

    def handle_equip_upgrade(self):
        if self.squad_tree_selection is None or treewidget_item_is_top_level(self.squad_tree_selection):
            return
        if self.upgrade_name_selected is None:
            logging.info("No upgrade selected - select an upgrade and try again.")
        pilot_item = self.squad_tree_selection.parent()
        pilot_data = self.squad.get_pilot_data(pilot_item)
        self.equip_upgrade(self.upgrade_name_selected, pilot_data, self.squad_tree_selection.parent(), self.squad_tree_selection)

    def equip_upgrade(self, upgrade_name: str, pilot_data: PilotEquip, parent_item=None, select_item=None):
        upgrade_dict = self.upgrades.get_upgrade(upgrade_name)
        upgrade_cost = Upgrades.get_filtered_upgrade_cost(
            upgrade_dict, pilot_data)
        upgrade_slots = Upgrades.get_upgrade_slots(upgrade_dict)
        equipped = pilot_data.equip_upgrade(
            upgrade_slots, upgrade_name, upgrade_cost, upgrade_dict)
        if equipped:
            self.refresh_squad_upgrade_slots(parent_select_item=parent_item, select_item=select_item)
        self.handle_squad_click()
        self.update_costs()
        self.viewer.populate_squad_viewer(self.squad)

    def unequip_upgrade(self):
        if self.squad_tree_selection is None or treewidget_item_is_top_level(self.squad_tree_selection):
            return
        pilot_item = self.squad_tree_selection.parent()
        pilot_data = self.squad.get_pilot_data(pilot_item)
        # Something is equipped
        if len(self.squad_tree_upgrade_name_selection) > 0:
            unequipped = pilot_data.unequip_upgrade(
                self.squad_tree_upgrade_name_selection)
            if unequipped:
                self.refresh_squad_upgrade_slots(self.squad_tree_selection.parent(), self.squad_tree_selection)
        self.handle_squad_click()
        self.update_costs()
        self.viewer.populate_squad_viewer(self.squad)

    def update_costs(self):
        """updates the UI cost labels based on squad list"""
        total_pilot = 0
        total_upgrade = 0
        for _, v in self.squad.squad_dict.items():
            total_pilot += v.cost
            total_upgrade += v.total_equipped_upgrade_cost
        total_cost = total_pilot + total_upgrade
        self.ui.total_pilot_cost_label.setText(str(total_pilot))
        self.ui.total_upgrade_cost.setText(str(total_upgrade))
        self.ui.total_cost_label.setText(str(total_cost))

    @property
    def squad_tree_bottom_index(self):
        return self.ui.squad_tree_widget.topLevelItemCount()

    @property
    def squad_tree_selection(self) -> QtWidgets.QTreeWidgetItem:
        try:
            val = self.ui.squad_tree_widget.selectedItems()[0]
        except IndexError:
            return
        return val

    def handle_new_upgrade_data(self, data):
        insert_flag = self.definition_form.insert_new_upgrade_entry(data)
        if insert_flag:
            self.upgrade_form.show()
        self.reload_data()

    def update_faction(self):
        if self.faction_selected is None:
            return
        self.ui.ship_list_widget.clear()
        self.ui.pilot_list_widget.clear()
        self.ui.upgrade_list_widget.clear()
        self.ui.pilot_keyword_label.clear()
        faction_name = self.faction_selected
        faction = self.xwing.get_faction(faction_name)
        populate_list_widget(faction.ship_names_for_gui,
                             self.ui.ship_list_widget, self.ship_icons_dir)

    def update_ship(self):
        self.ui.upgrade_list_widget.blockSignals(True)
        self.ui.upgrade_list_widget.clear()
        self.ui.upgrade_list_widget.blockSignals(False)
        self.ui.pilot_list_widget.blockSignals(True)
        self.pilot_image_label = None
        self.ui.main_card_viewer.add_card(None)
        ship_name = self.ship_selected_encoded
        ship = self.xwing.get_ship(self.faction_selected, ship_name)
        if ship is None:
            return
        self.ui.ship_name_label.setText(prettify_name(ship_name))
        self.ui.base_label.setText(prettify_name(ship.base))
        self.ui.initiative_label.setText(
            str(ship.initiative_list).replace('[', '').replace(']', ''))
        low, high = ship.point_range
        self.ui.points_label.setText(f"{low} - {high}")
        self.ui.maneuver_image_label.setPixmap(
            image_path_to_qpixmap(self.maneuvers_dir / f"{ship_name}.png"))
        update_action_layout(self.ui.ship_action_layout,
                             ship.actions, self.actions_dir)
        update_action_layout(self.ui.pilot_action_layout, [], self.actions_dir)
        update_upgrade_slot_layout(
            self.ui.ship_upgrade_slot_layout, ship.upgrade_slots, self.upgrade_slots_dir)
        update_upgrade_slot_layout(
            self.ui.pilot_upgrade_slot_layout, [], self.upgrade_slots_dir)

        self.ui.pilot_keyword_label.clear()

        # Populate the pilot list
        self.ui.pilot_list_widget.clear()
        self.ui.pilot_list_widget.blockSignals(False)
        populate_list_widget(ship.pilot_names_for_gui,
                             self.ui.pilot_list_widget)

    @property
    def pilot_image_label(self):
        return self.ui.pilot_image_label

    @pilot_image_label.setter
    def pilot_image_label(self, pilot_name: str):
        self.ui.main_card_viewer.set_card(
            image_path_to_qpixmap(self.pilots_dir / f"{pilot_name}.jpg"))

    def update_pilot(self):
        self.ui.main_card_viewer.add_card(None)
        self.ui.upgrade_list_widget.blockSignals(True)
        self.ui.upgrade_list_widget.clear()
        self.ui.upgrade_list_widget.blockSignals(False)
        pilot_name = self.pilot_name_selected
        if pilot_name is None or self.ship_selected_encoded is None:
            return
        self.pilot_image_label = pilot_name
        pilot = self.xwing.get_pilot(
            self.faction_selected, self.ship_selected_encoded, pilot_name)
        if pilot is None:
            return
        update_action_layout(self.ui.pilot_action_layout,
                             pilot["actions"], self.actions_dir)
        update_upgrade_slot_layout(
            self.ui.pilot_upgrade_slot_layout, pilot["upgrade_slots"], self.upgrade_slots_dir)

        if pilot["keywords"]:
            keyword_string = ", ".join([prettify_name(keyword) for keyword in pilot["keywords"]])
        else:
            keyword_string = ""

        self.ui.pilot_keyword_label.setText(keyword_string)

        # populate_list_widget(
        #     self.upgrades.all_upgrades_for_gui, self.ui.upgrade_list_widget)

    def update_upgrade(self):
        upgrade_name = self.upgrade_name_selected
        if upgrade_name is None:
            return
        self.ui.main_card_viewer.add_card(
            image_path_to_qpixmap(self.upgrades_dir / f"{upgrade_name}.jpg"))

    @property
    def faction_selected(self) -> str:
        try:
            val = self.ui.faction_list_widget.selectedItems()[0].text().lower()
        except IndexError:
            return
        return val

    @property
    def faction_item_selected(self) -> QtWidgets.QListWidgetItem:
        try:
            val = self.ui.faction_list_widget.selectedItems()[0]
        except IndexError:
            logging.info(
                "No faction selected - select a faction and try again.")
            return
        return val

    @property
    def ship_selected_encoded(self) -> str:
        try:
            val = gui_text_encode(
                self.ui.ship_list_widget.selectedItems()[0].text())
        except IndexError:
            return
        return val

    @property
    def pilot_selected_encoded(self) -> str:
        return gui_text_encode(self.pilot_selected_decoded)

    @property
    def pilot_selected_decoded(self) -> str:
        try:
            val = self.ui.pilot_list_widget.selectedItems()[0].text()
        except IndexError:
            return
        return val

    @property
    def pilot_name_selected(self) -> str:
        """returns the encoded lowercase name of the selected pilot from the pilot list widget."""
        try:
            val = get_pilot_name_from_list_item_text(
                self.ui.pilot_list_widget.selectedItems()[0].text())
        except IndexError:
            return
        return val

    @property
    def upgrade_name_selected(self) -> str:
        """returns the encoded lowercase name of the selected upgrade from the upgrade list widget."""

        try:
            val = get_upgrade_name_from_list_item_text(
                self.ui.upgrade_list_widget.selectedItems()[0].text())
        except IndexError:
            return
        return val

    @property
    def squad_tree_upgrade_name_selection(self) -> str:
        """returns the encoded lowercase name of the selected upgrade from the squad tree widget."""

        try:
            val = get_upgrade_name_from_list_item_text(
                self.squad_tree_selection.text(1))
        except IndexError:
            logging.info(
                "No squad tree item selected - select an item from the squad tree and try again.")
            return
        return val

    @property
    def data_dir(self) -> Path:
        return Path(__file__).parents[1] / "data"

    @property
    def factions_dir(self) -> Path:
        return Path(__file__).parents[1] / "data" / "resources" / "factions"

    @property
    def maneuvers_dir(self) -> Path:
        return Path(__file__).parents[1] / "data" / "resources" / "maneuvers"

    @property
    def pilots_dir(self) -> Path:
        return Path(__file__).parents[1] / "data" / "resources" / "pilots"

    @property
    def ship_icons_dir(self) -> Path:
        return Path(__file__).parents[1] / "data" / "resources" / "ship_icons"

    @property
    def upgrade_slots_dir(self) -> Path:
        return Path(__file__).parents[1] / "data" / "resources" / "upgrade_slots"

    @property
    def upgrades_dir(self) -> Path:
        return Path(__file__).parents[1] / "data" / "resources" / "upgrades"

    @property
    def actions_dir(self) -> Path:
        return Path(__file__).parents[1] / "data" / "resources" / "actions"

    def check_theme(self):
        app = QtWidgets.QApplication.instance()
        if self.settings_window.ui.theme_combo_box.currentText() == Settings.Theme.DARK.value:
            app.setStyle("Fusion")
            app.setPalette(DarkPalette())
            for widget, icon_path in self.widgets_with_icons.items():
                im = QtGui.QImage(icon_path)
                im.invertPixels()
                pixmap = QtGui.QPixmap.fromImage(im)
                widget.setIcon(pixmap)

        else:
            app.setStyle("windowsvista")
            app.setPalette(QtGui.QPalette())
            for widget, icon_path in self.widgets_with_icons.items():
                widget.setIcon(QtGui.QPixmap(icon_path))

    def configure_logging(self):
        self.ui.logging_plain_text_edit.setReadOnly(True)

        # Setup logging
        logfile = self.settings.log_file_dir / f"{self.application_name}.log"
        if logfile.exists():
            log_file_size = logfile.stat().st_size
            log_file_size_limit = 100 * 10**6  # size limit for log file in bytes
            if log_file_size > log_file_size_limit:
                os.remove(logfile)
        if not self.settings.log_file_dir.exists():
            os.makedirs(self.settings.log_file_dir)
        log_handler = RootLoggerHandler(filename=logfile)
        log_handler.sigLog.signal.connect(
            self.ui.logging_plain_text_edit.appendPlainText)

    def handle_show_settings_window(self):
        self.settings_window.show()

    @property
    def ready_for_export(self):
        if len(self.squad.squad_dict.items()) > 0:
            return True
        return False

    @property
    def squad_name(self) -> str:
        return self.ui.squad_name_line_edit.text()

    def export_excel(self):
        if self.ready_for_export:
            options = QtWidgets.QFileDialog.Options()
            filename, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, "Save As", "", "Excel Workbook (*.xlsx)", options=options
            )
            if filename:
                worker = Worker(self.squad.export_squad_as_excel, filename, self.squad_name)
                self.threadpool.start(worker)
        else:
            logging.info("Equip a pilot before trying to export your squad.")

    def import_excel(self):
        # TODO: Add checks for proper formatting (so you can't try to import any excel sheet)
        # TODO: This will only work for standard or epic mode
        if self.ready_for_export:
            buttonReply = QtWidgets.QMessageBox.question(
                self, "Warning", "Squad already in progress.  Are you sure you want to import?", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.Cancel)
            if buttonReply == QtWidgets.QMessageBox.Cancel:
                return
        options = QtWidgets.QFileDialog.Options()
        filename, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Import Squad", "", "Excel Workbook (*.xlsx)", options=options
        )
        if not filename:
            return
        self.squad = Squad()
        self.ui.squad_tree_widget.clear()

        workbook = load_workbook(filename)
        worksheet = workbook.active
        squad_name = worksheet['B1'].value
        faction_name = gui_text_encode(worksheet['B2'].value)
        self.ui.squad_name_line_edit.setText(squad_name)
        row_idx = 5
        while worksheet[f'A{row_idx}'].value is not None:
            pilot_name = gui_text_encode(worksheet[f'A{row_idx}'].value)
            ship_name = gui_text_encode(worksheet[f'B{row_idx}'].value)
            self.equip_pilot(faction_name, ship_name, pilot_name)
            col_idx = 5
            while worksheet[f'{get_column_letter(col_idx)}{row_idx}'].value is not None:
                upgrade_val = gui_text_encode(worksheet[f'{get_column_letter(col_idx)}{row_idx}'].value)
                pilot_data = self.squad.get_pilot_data_from_name(pilot_name)
                self.equip_upgrade(upgrade_val, pilot_data)
                col_idx += 1

            row_idx += 1
        for i in range(self.ui.faction_list_widget.count()):
            item = self.ui.faction_list_widget.item(i)
            if gui_text_encode(item.text()) == faction_name:
                item.setSelected(True)

    def closeEvent(self, event):
        buttonReply = QtWidgets.QMessageBox.question(
            self, "Warning", "Are you sure you want to quit?", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.Cancel)
        if buttonReply == QtWidgets.QMessageBox.Yes:
            # Add closing behaviors here
            self.definition_form.close()
            self.upgrade_form.close()
            self.viewer.close()
            event.accept()
        else:
            event.ignore()
